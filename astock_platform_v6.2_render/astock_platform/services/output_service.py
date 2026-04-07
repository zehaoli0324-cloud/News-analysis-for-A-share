"""
输出服务：终端打印 & Excel 导出
"""
import os
from datetime import datetime
from typing import List

from config.settings import (
    green, yellow, cyan, bold, dim, red,
)
from models.data_model import ensure_dict


# ─────────────────────────────────────────────
#  单位格式化工具函数
# ─────────────────────────────────────────────

def format_amount(value, unit="万元"):
    """格式化金额，自动转换单位
    
    例如：
    - 13497195.9 万元 → 1.35 亿元
    - 38629271969.0 元 → 386.29 亿元
    """
    if value is None or value == "" or value == "nan":
        return "未知"
    
    try:
        # 尝试转换为数字
        num = float(str(value).replace(",", ""))
        
        # 根据大小自动选择单位
        if unit == "元":
            if abs(num) >= 1e8:
                return f"{num/1e8:.2f}亿元"
            elif abs(num) >= 1e4:
                return f"{num/1e4:.2f}万元"
            else:
                return f"{num:.2f}元"
        elif unit == "万元":
            if abs(num) >= 10000:
                return f"{num/10000:.2f}亿元"
            else:
                return f"{num:.2f}万元"
        elif unit == "亿元":
            return f"{num:.2f}亿元"
        else:
            return f"{num:.2f}{unit}"
    except:
        return str(value)


def format_volume(value):
    """格式化成交量/成交额"""
    return format_amount(value, "元")


def format_money_wan(value):
    """格式化万元为单位的数据"""
    return format_amount(value, "万元")


# ─────────────────────────────────────────────
#  终端打印（新模板结构）
# ─────────────────────────────────────────────

def print_news(symbol, meta, stock_news, macro_news, caixin_news,
               macro_events, holder_news, company_news, sector_data,
               monetary=None, movement_news=None, spot_info=None, lhb_data=None):
    """格式化打印所有采集数据到终端（新模板结构）"""
    name       = meta.get("name", "") or symbol
    industry   = meta.get("industry", "未知")
    keywords   = meta.get("keywords", [])
    sector_data = sector_data or {}
    now = datetime.now().strftime("%Y-%m-%d")

    print(bold(f"\n{'═'*80}"))
    print(bold(f"  {name}（{symbol}）| {industry} | 分析日期：{now}"))
    print(bold(f"{'═'*80}"))

    # 一、数据质量声明
    print(bold(cyan("\n一、数据质量声明（必须先读）")))
    print_data_quality_declaration(
        symbol, company_news, stock_news, sector_data,
        holder_news, lhb_data, movement_news
    )

    # 二、公司基本面速览
    print(bold(cyan("\n二、公司基本面速览（仅基于可验证信息）")))
    print_company_basics(symbol, name, industry, holder_news, stock_news, lhb_data)

    # 三、板块与行业环境
    print(bold(cyan("\n三、板块与行业环境（可用数据充分）")))
    print_sector_environment(sector_data, movement_news, industry, meta)

    # 四、资金面与龙虎榜席位分析
    print(bold(cyan("\n四、资金面与龙虎榜席位分析（关键数据）")))
    print_lhb_analysis(lhb_data, sector_data, name)

    # 五、大股东信号
    print(bold(cyan("\n五、大股东信号")))
    print_holder_signals(holder_news)

    # 六、社会舆情与产业链（国内多源）
    print(bold(cyan("\n六、社会舆情与产业链（多维度资讯搜索）")))
    print_social_and_chain_news(company_news, sector_data)

    # 七、货币政策（弱相关）
    print(bold(cyan("\n七、货币政策（弱相关）")))
    print_monetary(monetary, industry=industry)

    # 合计
    mon_count = len(monetary.get("lpr", [])) + len(monetary.get("rrr", [])) if monetary else 0
    spot = sector_data.get("spot", {})
    sector_news = [ensure_dict(n) for n in sector_data.get("sector_news", [])]
    chain_news  = sector_data.get("chain_news", [])
    total = (len(company_news) + len(stock_news) + len(sector_news)
             + len(macro_news) + len(caixin_news)
             + len(macro_events) + len(holder_news) + mon_count)

    print(bold(f"\n{'─'*80}"))
    print(dim(f"  合计 {total} 条数据，即将进行 AI 分析..."))
    print(bold(f"{'─'*80}"))


def print_data_quality_declaration(symbol, company_news, stock_news, sector_data,
                                    holder_news, lhb_data, movement_news):
    """打印数据质量声明"""
    print(bold(f"  {'数据类别':<15}  {'状态':<6}  {'说明'}"))
    print(bold(f"  {'─'*15}  {'─'*6}  {'─'*50}"))

    # 公司专项公告（区分正式公告和GNews舆情）
    formal_announcements = [n for n in company_news if n.get('type') != 'social']
    social_news = [n for n in company_news if n.get('type') == 'social']
    
    if formal_announcements:
        print(f"  {'公司专项公告':<15}  {yellow('⚠️ 有限'):<6}  正式公告{len(formal_announcements)}条，可能含其他公司")
    else:
        print(f"  {'公司专项公告':<15}  {red('❌ 缺失'):<6}  {symbol}业绩预告、诉讼等正式公告缺失")

    # 个股新闻
    status_stock = green("✅ 可用") if stock_news else red("❌ 缺失")
    desc_stock = f"个股新闻{len(stock_news)}条（龙虎榜新闻、行业资金流出榜）" if stock_news else "无个股新闻"
    print(f"  {'个股新闻':<15}  {status_stock:<6}  {desc_stock}")

    # 板块数据
    spot = sector_data.get("spot", {})
    ff = sector_data.get("fund_flow", {})
    sector_news = [ensure_dict(n) for n in sector_data.get("sector_news", [])]
    status_sector = green("✅ 可用") if (spot or ff or sector_news) else red("❌ 缺失")
    desc_sector = ""
    if spot:
        chg = spot.get("涨跌幅", spot.get("今日涨跌幅", "?"))
        vol = spot.get("成交额", "?")
        desc_sector += f"板块涨跌{chg}%，成交额{vol}"
    if ff:
        inflow = ff.get("inflow", "")
        desc_sector += f"，主力净流出{inflow}"
    if sector_news:
        desc_sector += f"，行业新闻{len(sector_news)}条"
    desc_sector = desc_sector or "无板块数据"
    print(f"  {'板块数据':<15}  {status_sector:<6}  {desc_sector}")

    # 产业链新闻（GNews）
    chain_news = sector_data.get("chain_news", []) if sector_data else []
    if chain_news:
        upstream_count = len([n for n in chain_news if n.get('chain') == 'upstream'])
        downstream_count = len([n for n in chain_news if n.get('chain') == 'downstream'])
        policy_count = len([n for n in chain_news if n.get('chain') == 'policy'])
        print(f"  {'产业链新闻':<15}  {green('✅ 可用'):<6}  上游{upstream_count}条/下游{downstream_count}条/政策{policy_count}条")
    else:
        print(f"  {'产业链新闻':<15}  {red('❌ 缺失'):<6}  无产业链数据（上下游动态缺失）")

    # 龙虎榜席位
    lhb_data = lhb_data or {}
    has_lhb = lhb_data.get("has_lhb", False)
    has_details = len(lhb_data.get("details", [])) > 0
    status_lhb = green("✅ 可用") if (has_lhb and has_details) else (green("✅ 可用") if has_lhb else red("❌ 缺失"))
    desc_lhb = ""
    if has_lhb and has_details:
        details = lhb_data.get("details", [])
        inst_count = sum(1 for d in details if "机构" in str(d.get("seat", "")))
        retail_count = sum(1 for d in details if "拉萨" in str(d.get("seat", "")))
        desc_lhb = f"明细完整，机构{inst_count}家，散户{retail_count}家"
    elif has_lhb:
        desc_lhb = "有龙虎榜，但无席位明细"
    else:
        desc_lhb = "无龙虎榜数据"
    print(f"  {'龙虎榜席位':<15}  {status_lhb:<6}  {desc_lhb}")

    # 大股东持股
    status_holder = green("✅ 可用") if holder_news else red("❌ 缺失")
    desc_holder = f"大股东{len(holder_news)}条" if holder_news else "无大股东数据"
    if holder_news and len(holder_news) > 0:
        first = ensure_dict(holder_news[0])
        ratio = first.get("ratio", "")
        if "5" in str(ratio) or "6" in str(ratio) or "7" in str(ratio):
            desc_holder += f"，{first.get('holder','')}绝对控股{ratio}"
    print(f"  {'大股东持股':<15}  {status_holder:<6}  {desc_holder}")

    # 社会舆情（GNews）
    social_news = [n for n in company_news if n.get('type') == 'social']
    if social_news:
        print(f"  {'社会舆情':<15}  {green('✅ 可用'):<6}  搜索到{len(social_news)}条舆情")
    else:
        print(f"  {'社会舆情':<15}  {red('❌ 缺失'):<6}  无社会舆情数据")

    # 同行对比（GNews竞品分析）
    comp_report = sector_data.get('competitor_report', '') if sector_data else ''
    if comp_report:
        print(f"  {'同行对比':<15}  {green('✅ 可用'):<6}  已生成竞品对比报告")
    else:
        print(f"  {'同行对比':<15}  {red('❌ 缺失'):<6}  无竞品对比数据（keyword_dict 竞争对手为空或 GNews 失败）")

    print(dim(f"\n  分析原则：只基于有效数据推理；对缺失关键信息明确提示'无法判断，建议人工核查'。"))


def print_company_basics(symbol, name, industry, holder_news, stock_news, lhb_data):
    """打印公司基本面速览"""
    print(bold(f"  {'项目':<15}  {'内容':<40}  {'来源/备注'}"))
    print(bold(f"  {'─'*15}  {'─'*40}  {'─'*30}"))

    # 主营业务：用行业名代替硬编码
    print(f"  {'主营业务':<15}  {industry:<40}  {'公开信息，非本次采集'}")

    # 当前股价表现
    lhb_data = lhb_data or {}
    has_lhb = lhb_data.get("has_lhb", False)
    if has_lhb:
        concl = lhb_data.get("conclusion", "")
        print(f"  {'当前股价表现':<15}  {concl:<40}  {'东方财富个股新闻'}")
    else:
        print(f"  {'当前股价表现':<15}  {'无龙虎榜数据':<40}  {'无'}")

    # 大股东结构
    if holder_news:
        holder_str = ""
        for i, h_raw in enumerate(holder_news[:3], 1):
            h = ensure_dict(h_raw)
            holder_str += f"{h.get('holder','')}{h.get('ratio','')}；"
        if len(holder_news) > 3:
            holder_str += "等"
        print(f"  {'大股东结构':<15}  {holder_str[:40]:<40}  {'大股东持股数据'}")
    else:
        print(f"  {'大股东结构':<15}  {'无数据':<40}  {'无'}")

    # 近期自身公告
    print(f"  {'近期自身公告':<15}  {yellow('⚠️ 无数据'):<40}  采集失败，需手动访问巨潮资讯（{symbol}）核查业绩预告、诉讼等")

    # 关键风险
    print(red(f"\n  🚨 关键风险：{symbol}自身的2025年度业绩预告、重大诉讼进展等核心信息缺失，以下分析无法覆盖基本面的重大变化。"))


def print_sector_environment(sector_data, movement_news, industry, meta=None):
    """打印板块与行业环境"""
    spot = sector_data.get("spot", {})
    ff = sector_data.get("fund_flow", {})
    sector_news = [ensure_dict(n) for n in sector_data.get("sector_news", [])]
    movement_news = [ensure_dict(n) for n in (movement_news or [])]
    name = (meta or {}).get("name", "")
    # 竞品列表用于事件过滤，不再硬编码
    competitors_kws = (meta or {}).get("keyword_dict", {}).get("竞争对手", [])

    # 3.1 板块当日表现
    print(bold(f"\n  3.1 板块当日表现"))
    if spot:
        chg = spot.get("涨跌幅", spot.get("今日涨跌幅", "?"))
        vol = spot.get("成交额", "?")
        clr = green if "-" not in str(chg) else red
        print(f"    {industry}板块整体{clr(f'涨跌 {chg}%')}，成交额{vol}。")
    if ff:
        inflow = ff.get("inflow", "")
        clr = green if inflow and "-" not in str(inflow) else red
        print(f"    主力{clr(f'净流向 {inflow}')}（占比{ff.get('ratio','')}%）。")

    # 3.2 行业事件驱动分析
    print(bold(f"\n  3.2 行业事件驱动分析（非噪音）"))
    col4_header = f"对{name}的影响传导" if name else "影响传导"
    print(bold(f"    {'事件':<35}  {'日期':<12}  {'信号解读':<20}  {col4_header}"))
    print(bold(f"    {'─'*35}  {'─'*12}  {'─'*20}  {'─'*30}"))

    all_events = []
    for n in sector_news + movement_news:
        title = n.get("title", "")
        t     = n.get("time", n.get("date", ""))[:10]
        # 排除自身条目，纳入同行/竞品相关事件
        if name and name in title:
            continue
        if competitors_kws and any(c in title for c in competitors_kws):
            all_events.append((title, t))
        elif not competitors_kws:
            # 无竞品词时显示全部板块事件
            all_events.append((title, t))

    for title, t in all_events[:5]:
        interpretation = "中性"
        impact = "需结合个股判断"
        if "澄清" in title and ("AI" in title or "投资" in title):
            interpretation = "概念退潮"
            impact = "板块情绪受牵连"
        elif "转让" in title and "股权" in title:
            interpretation = "行业收缩"
            impact = "同业前景承压"
        elif "分红" in title or "回购" in title:
            interpretation = "正常经营"
            impact = "无直接关联"
        elif "利好" in title or "政策" in title:
            interpretation = "政策利好"
            impact = "板块整体受益"
        print(f"    {title[:34]:<35}  {t:<12}  {interpretation:<20}  {impact}")

    if not all_events:
        print(dim("    暂无同行事件，板块驱动因素不明确。"))

    print(dim(f"\n    板块判断：基于以上数据，{industry}板块近期动态需结合个股基本面综合判断。"))


def print_lhb_analysis(lhb_data, sector_data, name=""):
    """打印资金面与龙虎榜席位分析"""
    lhb_data = lhb_data or {}
    spot = sector_data.get("spot", {})
    ff = sector_data.get("fund_flow", {})
    details = lhb_data.get("details", [])

    # 4.1 整体资金流向
    print(bold(f"\n  4.1 整体资金流向"))
    if ff:
        inflow = ff.get("inflow", "")
        clr = green if inflow and "-" not in str(inflow) else red
        print(f"    板块主力{clr(f'净流向 {inflow}')}。")
    if lhb_data.get("has_lhb"):
        concl = lhb_data.get("conclusion", "")
        label = name if name else "个股"
        print(f"    {label}自身龙虎榜：{concl}。")

    # 4.2 席位明细
    print(bold(f"\n  4.2 席位明细"))
    if details:
        print(bold(f"    {'席位类型':<15}  {'营业部名称':<20}  {'风格/行为':<20}  {'信号'}"))
        print(bold(f"    {'─'*15}  {'─'*20}  {'─'*20}  {'─'*30}"))

        for d in details[:5]:
            seat = d.get("seat", "未知")
            style = d.get("style", "未知")
            signal = d.get("signal", "中性")
            seat_type = "其他"
            if "机构" in seat:
                seat_type = "机构专用"
            elif "拉萨" in seat:
                seat_type = "散户大本营"

            print(f"    {seat_type:<15}  {seat[:19]:<20}  {style[:19]:<20}  {signal}")


        print(dim(f"\n    当前可推断："))
        retail_count = sum(1 for d in details if "拉萨" in str(d.get("seat", "")))
        inst_count = sum(1 for d in details if "机构" in str(d.get("seat", "")))
        if retail_count > 0:
            print(dim(f"    - 散户席位密集出现 → 短期抛压主要来自散户恐慌盘。"))
        if inst_count > 0:
            print(dim(f"    - 机构席位出现 → 存在专业资金关注，但方向不明。"))
    else:
        print(dim("    无席位明细数据。"))


def print_holder_signals(holder_news):
    """打印大股东信号"""
    if not holder_news:
        print(dim("  无大股东数据。"))
        return

    first = ensure_dict(holder_news[0])
    holder = first.get("holder", "")
    ratio = first.get("ratio", "")
    print(f"  {holder}持股 {ratio}，处于绝对控股地位。")
    print(green("  正面：控制权稳定，决策效率高。"))
    print(red("  风险：一旦减持，抛压极大；需核查股权质押比例（本次无数据）。"))

    # 外资持股
    foreign = []
    for h_raw in holder_news:
        h = ensure_dict(h_raw)
        name = h.get("holder", "")
        if "BARCLAYS" in name or "J.P" in name or "MORGAN" in name:
            foreign.append(name)
    if foreign:
        print(dim(f"  前十大股东中出现 {', '.join(foreign[:2])} 等外资，虽然持股均较低，但说明有外资关注。"))

    print(dim(f"\n  结论：大股东结构无明显异常，但绝对控股的减持风险需持续监控。"))


def print_social_and_chain_news(company_news, sector_data):
    """打印社会舆情与产业链（国内多源）"""
    
    # 社会舆情（来自GNews）
    social_news = [n for n in company_news if n.get('type') == 'social']
    if social_news:
        print(f"\n  6.1 社会舆情（{len(social_news)} 条）")
        for i, news in enumerate(social_news, 1):
            t    = dim(news.get('time', '')[:16])
            src  = dim(f"[{news.get('source','')}]")
            title = news.get('title', '')
            print(f"    {i:>2}. [{t}] {title}  {src}")
    else:
        print(f"\n  6.1 社会舆情：{red('❌ 无数据')}")
    
    # 产业链新闻（来自多源）
    chain_news = sector_data.get('chain_news', []) if sector_data else []
    if chain_news:
        print(f"\n  6.2 产业链新闻（多源搜索）")
        upstream_news   = [n for n in chain_news if n.get('chain') == 'upstream']
        downstream_news = [n for n in chain_news if n.get('chain') == 'downstream']
        policy_news_c   = [n for n in chain_news if n.get('chain') == 'policy']

        if upstream_news:
            print(f"    上游新闻（{green(str(len(upstream_news)))}条）：")
            for news in upstream_news:
                print(f"      - [{dim(news.get('time','')[:10])}] {news.get('title', '')}")

        if downstream_news:
            print(f"    下游新闻（{green(str(len(downstream_news)))}条）：")
            for news in downstream_news:
                print(f"      - [{dim(news.get('time','')[:10])}] {news.get('title', '')}")

        if policy_news_c:
            print(f"    政策新闻（{green(str(len(policy_news_c)))}条）：")
            for news in policy_news_c:
                print(f"      - [{dim(news.get('time','')[:10])}] {news.get('title', '')}")
    else:
        print(f"\n  6.2 产业链新闻：{red('❌ 无数据')}")
    
    # 竞品对比（国内多源搜索）
    comp_report = sector_data.get('competitor_report', '') if sector_data else ''
    if comp_report:
        print(f"\n  6.3 竞品对比分析")
        # 显示完整的竞品分析报告
        report_lines = comp_report.split('\n')
        for line in report_lines:
            if line.strip():
                print(f"  {line}")
    else:
        print(f"\n  6.3 竞品对比：{red('❌ 无数据')}")


def print_missing_sections(name="", industry=""):
    """打印社会舆情与产业链（完全缺失时的提示）"""
    print(bold(f"  {'维度':<15}  {'状态':<6}  {'建议补充方式'}"))
    print(bold(f"  {'─'*15}  {'─'*6}  {'─'*50}"))
    label = name if name else "该股"
    ind   = industry if industry else "所在行业"
    print(f"  {'社会舆情':<15}  {red('❌ 无'):<6}  手动搜索「{label}」相关新闻或行业动态平台")
    print(f"  {'产业链上游':<15}  {red('❌ 无'):<6}  可关注{ind}上游供应商财报或行业研报")
    print(f"  {'产业链下游':<15}  {red('❌ 无'):<6}  可关注{ind}下游消费/需求端数据")
    print(f"  {'同行对比':<15}  {red('❌ 无'):<6}  需手动采集竞品当日涨跌、新闻、龙虎榜")
    print(dim(f"\n  由于上述数据缺失，无法判断{label}的股价变动是行业共性还是个股问题。"))


def print_monetary(monetary, industry=""):
    """打印货币政策"""
    monetary = monetary or {}
    lpr_data = monetary.get("lpr", [])
    rrr_data = monetary.get("rrr", [])

    if lpr_data:
        lpr1 = lpr_data[0].get("1年期", "")
        lpr5 = lpr_data[0].get("5年期", "")
        print(f"  LPR 1年期{lpr1}，5年期{lpr5}；存款准备金率大型9.5%。")
        ind_note = f"{industry}行业" if industry else "相关行业"
        print(dim(f"\n  宽松利率环境有利于{ind_note}降低资金成本，属于温和利好，但非短期主要矛盾。"))
    else:
        print(dim("  无货币政策数据。"))


# ─────────────────────────────────────────────
#  Excel 导出（保持原样，后续可优化）
# ─────────────────────────────────────────────

def export_excel(symbol, stock_news, macro_news, caixin_news,
                 macro_events, ai_analysis):
    """将新闻数据和AI分析结果导出为Excel文件"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        print(yellow("  ⚠ 未安装 openpyxl，跳过导出。请运行：pip install openpyxl"))
        return

    stock_news   = [ensure_dict(n) for n in (stock_news   or [])]
    macro_news   = [ensure_dict(n) for n in (macro_news   or [])]
    caixin_news  = [ensure_dict(n) for n in (caixin_news  or [])]
    macro_events = [ensure_dict(n) for n in (macro_events or [])]

    wb  = Workbook()
    now = datetime.now()
    now_str  = now.strftime("%Y-%m-%d %H:%M")
    filename = f"股票分析_{symbol}_{now.strftime('%Y%m%d_%H%M')}.xlsx"

    header_font  = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    header_fill  = PatternFill("solid", start_color="1F4E79")
    sub_fill     = PatternFill("solid", start_color="2E75B6")
    normal_font  = Font(name="Arial", size=10)
    wrap_align   = Alignment(horizontal="left", vertical="top", wrap_text=True)
    center_align = Alignment(horizontal="center", vertical="center")
    thin_border  = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"),  bottom=Side(style="thin"),
    )
    alt_fill = PatternFill("solid", start_color="DEEAF1")

    def set_header(ws, row, col, value, fill=None):
        cell = ws.cell(row=row, column=col, value=value)
        cell.font      = header_font
        cell.fill      = fill or header_fill
        cell.alignment = center_align
        cell.border    = thin_border
        return cell

    def set_cell(ws, row, col, value, fill=None):
        cell = ws.cell(row=row, column=col, value=str(value) if value else "")
        cell.font      = normal_font
        cell.alignment = wrap_align
        cell.border    = thin_border
        if fill:
            cell.fill = fill
        return cell

    ws1 = wb.active
    ws1.title = "AI分析报告"
    ws1.merge_cells("A1:B1")
    ws1["A1"].value     = f"A股智能分析报告 — {symbol}   生成时间：{now_str}"
    ws1["A1"].font      = Font(name="Arial", bold=True, size=13, color="FFFFFF")
    ws1["A1"].fill      = PatternFill("solid", start_color="1F4E79")
    ws1["A1"].alignment = center_align
    ws1.row_dimensions[1].height = 30
    set_header(ws1, 2, 1, "项目", sub_fill)
    set_header(ws1, 2, 2, "内容", sub_fill)
    ws1.column_dimensions["A"].width = 16
    ws1.column_dimensions["B"].width = 90
    items = [
        ("股票代码", symbol),
        ("分析时间", now_str),
        ("数据来源", "东方财富 / 新浪 / 百度 / 网易财经"),
        ("AI模型",   "智谱 GLM-4-Flash"),
        ("新闻总数", f"{len(stock_news)+len(macro_news)+len(caixin_news)+len(macro_events)} 条"),
        ("AI分析结果", ai_analysis),
    ]
    for i, (k, v) in enumerate(items, 3):
        fill = alt_fill if i % 2 == 0 else None
        set_cell(ws1, i, 1, k, fill)
        set_cell(ws1, i, 2, v, fill)
        ws1.row_dimensions[i].height = max(200, len(str(v)) // 2) if k == "AI分析结果" else 20

    ws2 = wb.create_sheet("个股新闻")
    ws2.merge_cells("A1:C1")
    ws2["A1"].value     = f"东方财富·个股新闻 — {symbol}"
    ws2["A1"].font      = Font(name="Arial", bold=True, size=12, color="FFFFFF")
    ws2["A1"].fill      = PatternFill("solid", start_color="375623")
    ws2["A1"].alignment = center_align
    ws2.row_dimensions[1].height = 25
    for col, (col_name, w) in enumerate(zip(["序号", "发布时间", "新闻标题"], [6, 20, 80]), 1):
        set_header(ws2, 2, col, col_name, PatternFill("solid", start_color="538135"))
        ws2.column_dimensions[get_column_letter(col)].width = w
    for i, n in enumerate(stock_news, 1):
        fill = alt_fill if i % 2 == 0 else None
        set_cell(ws2, i+2, 1, i, fill)
        set_cell(ws2, i+2, 2, n.get("time", ""), fill)
        set_cell(ws2, i+2, 3, n.get("title", ""), fill)
        ws2.row_dimensions[i+2].height = 18

    ws3 = wb.create_sheet("宏观快讯")
    ws3.merge_cells("A1:D1")
    ws3["A1"].value     = "财联社·全球宏观快讯"
    ws3["A1"].font      = Font(name="Arial", bold=True, size=12, color="FFFFFF")
    ws3["A1"].fill      = PatternFill("solid", start_color="7B3F00")
    ws3["A1"].alignment = center_align
    ws3.row_dimensions[1].height = 25
    for col, (col_name, w) in enumerate(zip(["序号", "发布时间", "标题", "内容摘要"], [6, 20, 50, 60]), 1):
        set_header(ws3, 2, col, col_name, PatternFill("solid", start_color="C55A11"))
        ws3.column_dimensions[get_column_letter(col)].width = w
    for i, n in enumerate(macro_news, 1):
        fill = alt_fill if i % 2 == 0 else None
        content = n.get("content", "")
        set_cell(ws3, i+2, 1, i, fill)
        set_cell(ws3, i+2, 2, n.get("time", ""), fill)
        set_cell(ws3, i+2, 3, n.get("title", ""), fill)
        set_cell(ws3, i+2, 4, content if content != "nan" else "", fill)
        ws3.row_dimensions[i+2].height = 30 if content and content != "nan" else 18

    output_path = os.path.join(os.getcwd(), filename)
    wb.save(output_path)
    print(green(f"\n  ✅ Excel 已导出：{output_path}"))
    return output_path
